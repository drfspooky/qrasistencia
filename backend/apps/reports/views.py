import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from django.http import HttpResponse, FileResponse
from django.utils import timezone
from django.db.models import Count, Avg, Q
from django.shortcuts import get_object_or_404
from django.contrib.auth import get_user_model
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions, status
from drf_spectacular.utils import extend_schema, OpenApiParameter, OpenApiTypes

from apps.users.permissions import IsAdmin, IsTeacher, IsAdminOrReadOnly
from apps.academics.models import Section, Enrollment
from apps.attendance.models import ClassSession, AttendanceRecord, AttendanceAlert
from apps.users.models import StudentProfile

import io
from datetime import timedelta

User = get_user_model()


class DailySummaryAPIView(APIView):
    permission_classes = [IsTeacher | IsAdmin]

    def get(self, request):
        today = timezone.now().date()
        user = request.user
        
        # Filter sessions by role
        sessions = ClassSession.objects.filter(date=today)
        if user.role == 'teacher':
            sessions = sessions.filter(section__teacher__user=user)
            
        total_sessions = sessions.count()
        total_active = sessions.filter(status='active').count()
        total_closed = sessions.filter(status='closed').count()
        
        # Attendance breakdown for today's sessions
        records = AttendanceRecord.objects.filter(session__in=sessions)
        
        breakdown = {
            'total_sessions': total_sessions,
            'active_sessions': total_active,
            'closed_sessions': total_closed,
            'attendances': {
                'presente': records.filter(status='presente').count(),
                'tardanza': records.filter(status='tardanza').count(),
                'falta': records.filter(status='falta').count(),
                'retiro_anticipado': records.filter(status='retiro_anticipado').count(),
                'justificado': records.filter(status='justificado').count(),
            }
        }
        
        return Response(breakdown)


class ByStudentReportAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        parameters=[
            OpenApiParameter('student_id', OpenApiTypes.INT, OpenApiParameter.QUERY, description='Student Profile ID')
        ]
    )
    def get(self, request):
        user = request.user
        student_id = request.query_params.get('student_id')
        
        if student_id:
            if user.role not in ['superadmin', 'admin', 'teacher']:
                return Response({'detail': 'No tienes permisos para ver reportes de otros alumnos'}, status=status.HTTP_403_FORBIDDEN)
            student = get_object_or_404(StudentProfile, id=student_id)
        else:
            if user.role != 'student':
                return Response({'detail': 'Debe especificar student_id en los parámetros'}, status=status.HTTP_400_BAD_REQUEST)
            student = getattr(user, 'student_profile', None)
            if not student:
                return Response({'detail': 'Perfil de estudiante no encontrado'}, status=status.HTTP_404_NOT_FOUND)
                
        # Find student enrollments
        enrollments = Enrollment.objects.filter(student=student).select_related('section__course', 'section__period')
        
        courses_report = []
        for enrollment in enrollments:
            section = enrollment.section
            sessions = ClassSession.objects.filter(section=section, status='closed')
            total_sessions = sessions.count()
            
            records = AttendanceRecord.objects.filter(student=student, session__in=sessions)
            present = records.filter(status='presente').count()
            late = records.filter(status='tardanza').count()
            absent = records.filter(status='falta').count()
            justified = records.filter(status='justificado').count()
            early_leave = records.filter(status='retiro_anticipado').count()
            
            attended = present + late + justified
            attendance_rate = 100.0 if total_sessions == 0 else (attended / total_sessions) * 100.0
            
            # Semáforo
            if attendance_rate >= 85.0:
                traffic_light = 'green'
            elif attendance_rate >= 70.0:
                traffic_light = 'yellow'
            else:
                traffic_light = 'red'
                
            courses_report.append({
                'section_id': section.id,
                'section_code': section.code,
                'course_name': section.course.name,
                'course_code': section.course.code,
                'period': section.period.name,
                'teacher': section.teacher.user.get_full_name(),
                'total_sessions': total_sessions,
                'present': present,
                'late': late,
                'absent': absent,
                'justified': justified,
                'early_leave': early_leave,
                'attendance_percentage': round(attendance_rate, 2),
                'traffic_light': traffic_light
            })
            
        return Response({
            'student_name': student.user.get_full_name(),
            'student_code': student.student_code,
            'courses': courses_report
        })


class ByCourseReportAPIView(APIView):
    permission_classes = [IsTeacher | IsAdmin]

    @extend_schema(
        parameters=[
            OpenApiParameter('section_id', OpenApiTypes.INT, OpenApiParameter.QUERY, description='Section ID')
        ]
    )
    def get(self, request):
        user = request.user
        section_id = request.query_params.get('section_id')
        
        queryset = Section.objects.all().select_related('course', 'period', 'teacher__user')
        if section_id:
            queryset = queryset.filter(id=section_id)
        if user.role == 'teacher':
            queryset = queryset.filter(teacher__user=user)
            
        report_data = []
        for section in queryset:
            # Enrolled students count
            total_students = Enrollment.objects.filter(section=section).count()
            # Closed sessions
            closed_sessions = ClassSession.objects.filter(section=section, status='closed').count()
            
            # Attendance rates
            total_records = AttendanceRecord.objects.filter(session__section=section, session__status='closed')
            total_records_count = total_records.count()
            
            if total_records_count > 0:
                attended = total_records.filter(status__in=['presente', 'tardanza', 'justificado']).count()
                average_attendance = (attended / total_records_count) * 100.0
            else:
                average_attendance = 100.0
                
            # Alertas semáforo
            alerts = AttendanceAlert.objects.filter(student__enrollments__section=section, academic_period=section.period)
            red_count = alerts.filter(status='red').count()
            yellow_count = alerts.filter(status='yellow').count()
            green_count = alerts.filter(status='green').count()
            
            report_data.append({
                'section_id': section.id,
                'course_name': section.course.name,
                'course_code': section.course.code,
                'section_code': section.code,
                'teacher': section.teacher.user.get_full_name(),
                'period': section.period.name,
                'total_students': total_students,
                'total_sessions': closed_sessions,
                'average_attendance': round(average_attendance, 2),
                'alerts': {
                    'green': green_count,
                    'yellow': yellow_count,
                    'red': red_count
                }
            })
            
        return Response(report_data)


class ExportPDFReportAPIView(APIView):
    permission_classes = [IsTeacher | IsAdmin]

    @extend_schema(
        parameters=[
            OpenApiParameter('section_id', OpenApiTypes.INT, OpenApiParameter.QUERY, description='Section ID', required=True)
        ]
    )
    def get(self, request):
        section_id = request.query_params.get('section_id')
        if not section_id:
            return Response({'detail': 'Parámetro section_id es obligatorio'}, status=status.HTTP_400_BAD_REQUEST)
            
        section = get_object_or_404(Section, id=section_id)
        
        # Verify row-level permission for teachers
        if request.user.role == 'teacher' and section.teacher.user != request.user:
            return Response({'detail': 'No tienes permisos sobre esta sección de clase'}, status=status.HTTP_403_FORBIDDEN)
            
        # Fetch enrollments
        enrollments = Enrollment.objects.filter(section=section).select_related('student__user').order_by('student__user__last_name', 'student__user__first_name')
        # Fetch sessions
        sessions = ClassSession.objects.filter(section=section, status='closed').order_by('date', 'start_time')
        total_sessions = sessions.count()
        
        # Build document buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        story = []
        
        # Setup styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=18,
            textColor=colors.HexColor('#1E293B'),
            spaceAfter=15
        )
        subtitle_style = ParagraphStyle(
            'ReportSub',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=11,
            textColor=colors.HexColor('#475569'),
            spaceAfter=5
        )
        
        # Title & Subheaders
        story.append(Paragraph(f"Reporte de Asistencia Académica", title_style))
        story.append(Paragraph(f"<b>Curso:</b> {section.course.name} ({section.course.code})", subtitle_style))
        story.append(Paragraph(f"<b>Sección:</b> {section.code} | <b>Periodo:</b> {section.period.name}", subtitle_style))
        story.append(Paragraph(f"<b>Docente:</b> {section.teacher.user.get_full_name()}", subtitle_style))
        story.append(Paragraph(f"<b>Fecha de Generación:</b> {timezone.now().strftime('%d/%m/%Y %H:%M')}", subtitle_style))
        story.append(Spacer(1, 15))
        
        # Data grid preparation
        data = [['Código', 'Alumno', 'Pres.', 'Tard.', 'Falt.', 'Just.', '% Asist.', 'Riesgo']]
        
        for enrollment in enrollments:
            student = enrollment.student
            records = AttendanceRecord.objects.filter(student=student, session__in=sessions)
            p = records.filter(status='presente').count()
            t = records.filter(status='tardanza').count()
            f = records.filter(status='falta').count()
            j = records.filter(status='justificado').count()
            
            attended = p + t + j
            rate = 100.0 if total_sessions == 0 else (attended / total_sessions) * 100.0
            
            if rate >= 85.0:
                risk = 'Verde'
            elif rate >= 70.0:
                risk = 'Amarillo'
            else:
                risk = 'Rojo'
                
            data.append([
                student.student_code,
                student.user.get_full_name(),
                str(p),
                str(t),
                str(f),
                str(j),
                f"{rate:.1f}%",
                risk
            ])
            
        # Draw table
        t = Table(data, colWidths=[70, 200, 45, 45, 45, 45, 60, 60])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8FAFC')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TEXTCOLOR', (7, 1), (7, -1), colors.HexColor('#EF4444')), # Default RED text for risk
        ]))
        
        # Dynamic row styling for risk colors
        for i in range(1, len(data)):
            risk_val = data[i][7]
            if risk_val == 'Verde':
                t.setStyle(TableStyle([('TEXTCOLOR', (7, i), (7, i), colors.HexColor('#10B981'))]))
            elif risk_val == 'Amarillo':
                t.setStyle(TableStyle([('TEXTCOLOR', (7, i), (7, i), colors.HexColor('#F59E0B'))]))
                
        story.append(t)
        doc.build(story)
        
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename=f"reporte_asistencia_{section.course.code}_{section.code}.pdf")


class ExportExcelReportAPIView(APIView):
    permission_classes = [IsTeacher | IsAdmin]

    @extend_schema(
        parameters=[
            OpenApiParameter('section_id', OpenApiTypes.INT, OpenApiParameter.QUERY, description='Section ID', required=True)
        ]
    )
    def get(self, request):
        section_id = request.query_params.get('section_id')
        if not section_id:
            return Response({'detail': 'Parámetro section_id es obligatorio'}, status=status.HTTP_400_BAD_REQUEST)
            
        section = get_object_or_404(Section, id=section_id)
        
        # Verify row-level permission for teachers
        if request.user.role == 'teacher' and section.teacher.user != request.user:
            return Response({'detail': 'No tienes permisos sobre esta sección de clase'}, status=status.HTTP_403_FORBIDDEN)
            
        # Fetch enrollments
        enrollments = Enrollment.objects.filter(section=section).select_related('student__user').order_by('student__user__last_name', 'student__user__first_name')
        # Fetch all closed sessions
        sessions = ClassSession.objects.filter(section=section, status='closed').order_by('date', 'start_time')
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Matriz de Asistencia"
        
        # Styles
        font_header = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        font_sub = Font(name='Arial', size=10, italic=True)
        font_bold = Font(name='Arial', size=10, bold=True)
        
        fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        fill_present = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # green
        fill_late = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")    # yellow
        fill_absent = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # red
        fill_justified = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid") # blue
        
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        
        # Header Info
        ws.cell(row=1, column=1, value="REPORTE DE ASISTENCIA").font = Font(name='Arial', size=14, bold=True)
        ws.cell(row=2, column=1, value=f"Curso: {section.course.name} ({section.course.code})").font = font_sub
        ws.cell(row=3, column=1, value=f"Sección: {section.code} | Docente: {section.teacher.user.get_full_name()}").font = font_sub
        ws.cell(row=4, column=1, value=f"Periodo: {section.period.name} | Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M')}").font = font_sub
        
        # Grid Headers (starting at row 6)
        ws.cell(row=6, column=1, value="Código").font = font_header
        ws.cell(row=6, column=1).fill = fill_header
        ws.cell(row=6, column=2, value="Alumno").font = font_header
        ws.cell(row=6, column=2).fill = fill_header
        
        # Add a column for each session date
        col_index = 3
        for session in sessions:
            date_str = session.date.strftime("%d/%m")
            cell = ws.cell(row=6, column=col_index, value=date_str)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center')
            col_index += 1
            
        # Summary headers
        ws.cell(row=6, column=col_index, value="P").font = font_header
        ws.cell(row=6, column=col_index).fill = fill_header
        ws.cell(row=6, column=col_index).alignment = Alignment(horizontal='center')
        
        ws.cell(row=6, column=col_index+1, value="T").font = font_header
        ws.cell(row=6, column=col_index+1).fill = fill_header
        ws.cell(row=6, column=col_index+1).alignment = Alignment(horizontal='center')
        
        ws.cell(row=6, column=col_index+2, value="F").font = font_header
        ws.cell(row=6, column=col_index+2).fill = fill_header
        ws.cell(row=6, column=col_index+2).alignment = Alignment(horizontal='center')
        
        ws.cell(row=6, column=col_index+3, value="J").font = font_header
        ws.cell(row=6, column=col_index+3).fill = fill_header
        ws.cell(row=6, column=col_index+3).alignment = Alignment(horizontal='center')
        
        ws.cell(row=6, column=col_index+4, value="% Asist.").font = font_header
        ws.cell(row=6, column=col_index+4).fill = fill_header
        ws.cell(row=6, column=col_index+4).alignment = Alignment(horizontal='center')
        
        # Grid Data
        row_index = 7
        for enrollment in enrollments:
            student = enrollment.student
            ws.cell(row=row_index, column=1, value=student.student_code).border = thin_border
            ws.cell(row=row_index, column=2, value=student.user.get_full_name()).border = thin_border
            
            p, t, f_count, j = 0, 0, 0, 0
            curr_col = 3
            
            for session in sessions:
                record = AttendanceRecord.objects.filter(student=student, session=session).first()
                status_char = "-"
                fill = None
                
                if record:
                    if record.status == 'presente':
                        status_char = "P"
                        p += 1
                        fill = fill_present
                    elif record.status == 'tardanza':
                        status_char = "T"
                        t += 1
                        fill = fill_late
                    elif record.status == 'falta':
                        status_char = "F"
                        f_count += 1
                        fill = fill_absent
                    elif record.status == 'justificado':
                        status_char = "J"
                        j += 1
                        fill = fill_justified
                        
                cell = ws.cell(row=row_index, column=curr_col, value=status_char)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                if fill:
                    cell.fill = fill
                curr_col += 1
                
            # Write summaries
            ws.cell(row=row_index, column=curr_col, value=p).border = thin_border
            ws.cell(row=row_index, column=curr_col).alignment = Alignment(horizontal='center')
            
            ws.cell(row=row_index, column=curr_col+1, value=t).border = thin_border
            ws.cell(row=row_index, column=curr_col+1).alignment = Alignment(horizontal='center')
            
            ws.cell(row=row_index, column=curr_col+2, value=f_count).border = thin_border
            ws.cell(row=row_index, column=curr_col+2).alignment = Alignment(horizontal='center')
            
            ws.cell(row=row_index, column=curr_col+3, value=j).border = thin_border
            ws.cell(row=row_index, column=curr_col+3).alignment = Alignment(horizontal='center')
            
            # Percentage calculation
            total_s = len(sessions)
            rate = 100.0 if total_s == 0 else ((p + t + j) / total_s) * 100.0
            cell_rate = ws.cell(row=row_index, column=curr_col+4, value=f"{rate:.1f}%")
            cell_rate.border = thin_border
            cell_rate.alignment = Alignment(horizontal='center')
            cell_rate.font = font_bold
            
            row_index += 1
            
        # Adjust columns width
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        
        # Save workbook to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer, 
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_asistencia_{section.course.code}_{section.code}.xlsx"'
        return response
